import os
import requests
from twilio.rest import Client
from openpyxl import Workbook, load_workbook
from datetime import datetime
from django.conf import settings


def _send_via_resend(to_email, subject, html_content, reply_to=None):
    """Send an email using the Resend HTTP API.

    This replaces smtplib which is blocked on Render's free tier.
    Resend uses HTTPS so it works on any hosting platform.
    """
    api_key = getattr(settings, 'RESEND_API_KEY', None)
    if not api_key:
        print("RESEND_API_KEY not configured — skipping email")
        return False

    # Use verified domain sender, or fallback to Resend's onboarding address
    from_email = getattr(settings, 'RESEND_FROM_EMAIL', 'Grovix Studio <onboarding@resend.dev>')

    payload = {
        "from": from_email,
        "to": [to_email],
        "subject": subject,
        "html": html_content,
    }
    if reply_to:
        payload["reply_to"] = reply_to

    response = requests.post(
        "https://api.resend.com/emails",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=10,
    )

    if response.status_code == 200:
        print(f"Email sent to {to_email}: {response.json()}")
        return True
    else:
        print(f"Resend error ({response.status_code}): {response.text}")
        return False


def send_email_to_client(inquiry):
    """Send email notification to client about new inquiry"""
    try:
        subject = f"New Inquiry - {inquiry.inquiry_id} - {inquiry.selected_plan}"
        html = f"""
        <html>
          <body style="font-family: Arial, sans-serif; background-color: #0a0a0f; color: #ffffff; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: linear-gradient(135deg, rgba(0, 217, 255, 0.1), rgba(168, 85, 247, 0.1)); border: 1px solid rgba(0, 217, 255, 0.3); border-radius: 20px; padding: 40px;">
              <h1 style="color: #00d9ff; text-align: center; margin-bottom: 30px;">🎯 New Customer Inquiry</h1>
              
              <div style="background: rgba(20, 20, 30, 0.8); padding: 25px; border-radius: 15px; margin-bottom: 20px;">
                <h2 style="color: #a855f7; margin-bottom: 20px;">Inquiry Details</h2>
                <p><strong style="color: #00d9ff;">Inquiry ID:</strong> {inquiry.inquiry_id}</p>
                <p><strong style="color: #00d9ff;">Name:</strong> {inquiry.name}</p>
                <p><strong style="color: #00d9ff;">Email:</strong> {inquiry.email}</p>
                <p><strong style="color: #00d9ff;">Phone:</strong> {inquiry.phone}</p>
                {f'<p><strong style="color: #00d9ff;">Company:</strong> {inquiry.company}</p>' if inquiry.company else ''}
                <p><strong style="color: #00d9ff;">Selected Plan:</strong> <span style="color: #fbbf24;">{inquiry.selected_plan}</span></p>
                <p><strong style="color: #00d9ff;">Date:</strong> {inquiry.created_at.strftime('%B %d, %Y at %I:%M %p')}</p>
              </div>
              
              {f"""<div style="background: rgba(20, 20, 30, 0.8); padding: 25px; border-radius: 15px; margin-bottom: 20px;">
                <h3 style="color: #a855f7; margin-bottom: 15px;">Message</h3>
                <p style="line-height: 1.6;">{inquiry.message}</p>
              </div>""" if inquiry.message else ''}
              
              <div style="text-align: center; margin-top: 30px;">
                <p style="color: rgba(255, 255, 255, 0.7);">Contact the customer as soon as possible!</p>
              </div>
            </div>
          </body>
        </html>
        """

        return _send_via_resend(
            to_email=settings.CLIENT_EMAIL,
            subject=subject,
            html_content=html,
            reply_to=inquiry.email,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Email error: {e}")
        return False


def send_email_to_customer(inquiry):
    """Send a confirmation email to the customer who submitted the inquiry."""
    if not inquiry.email:
        return False

    try:
        subject = f"Thank you for your inquiry — {inquiry.selected_plan}"
        html = f"""
        <html>
          <body style="font-family: Arial, sans-serif; background-color: #0a0a0f; color: #ffffff; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: linear-gradient(135deg, rgba(0, 217, 255, 0.1), rgba(168, 85, 247, 0.1)); border: 1px solid rgba(0, 217, 255, 0.3); border-radius: 20px; padding: 40px;">
              <h1 style="color: #00d9ff; text-align: center; margin-bottom: 30px;">✅ Inquiry Received</h1>

              <div style="background: rgba(20, 20, 30, 0.8); padding: 25px; border-radius: 15px; margin-bottom: 20px;">
                <h2 style="color: #a855f7; margin-bottom: 20px;">We got your message</h2>
                <p>Hi <strong style="color: #00d9ff;">{inquiry.name}</strong>,</p>
                <p>Thanks for reaching out about <strong style="color: #fbbf24;">{inquiry.selected_plan}</strong>. We've received your inquiry and will respond within 24 hours.</p>
                <p><strong>Inquiry ID:</strong> {inquiry.inquiry_id}</p>
                <p><strong>What you submitted:</strong></p>
                <ul>
                  <li><strong>Email:</strong> {inquiry.email}</li>
                  <li><strong>Phone:</strong> {inquiry.phone}</li>
                  {f'<li><strong>Company:</strong> {inquiry.company}</li>' if inquiry.company else ''}
                </ul>
              </div>

              {f"""<div style="background: rgba(20, 20, 30, 0.8); padding: 25px; border-radius: 15px; margin-bottom: 20px;">
                <h3 style="color: #a855f7; margin-bottom: 15px;">Your message</h3>
                <p style="line-height: 1.6;">{inquiry.message}</p>
              </div>""" if inquiry.message else ''}

              <div style="text-align: center; margin-top: 30px;">
                <p style="color: rgba(255, 255, 255, 0.7);">If you need to update your inquiry, just reply to this email.</p>
                <p style="color: rgba(255, 255, 255, 0.7);">Best,</p>
                <p style="color: rgba(255, 255, 255, 0.7);"><strong>Grovix Studio Team</strong></p>
              </div>
            </div>
          </body>
        </html>
        """

        return _send_via_resend(
            to_email=inquiry.email,
            subject=subject,
            html_content=html,
            reply_to=settings.CLIENT_EMAIL,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Email error (customer): {e}")
        return False


def send_whatsapp_message(inquiry):
    """Send a confirmation message to the customer via Twilio.

    This will attempt to send via WhatsApp if a WhatsApp sender number is configured.
    If WhatsApp fails, it will fall back to regular SMS.

    The customer's phone number is normalized to remove spaces/dashes so it works with Twilio.
    """
    if not inquiry.phone:
        print("SMS error: no phone number provided")
        return False

    def normalize_phone(phone: str) -> str:
        # Remove common punctuation and whitespace
        cleaned = phone.strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        # Ensure international format (starts with +)
        if cleaned and not cleaned.startswith('+'):
            cleaned = '+' + cleaned
        return cleaned

    customer_number = normalize_phone(inquiry.phone)

    try:
        client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)

        message_body = (
            f"Thank you {inquiry.name}! Your inquiry {inquiry.inquiry_id} for {inquiry.selected_plan} has been received. "
            "We'll contact you within 24 hours. - Grovix Studio"
        )

        # Prefer WhatsApp if configured
        if settings.TWILIO_WHATSAPP_NUMBER:
            whatsapp_from = settings.TWILIO_WHATSAPP_NUMBER
            if not whatsapp_from.startswith('whatsapp:'):
                whatsapp_from = f'whatsapp:{whatsapp_from}'

            whatsapp_to = customer_number
            if not whatsapp_to.startswith('whatsapp:'):
                whatsapp_to = f'whatsapp:{whatsapp_to}'

            message = client.messages.create(
                from_=whatsapp_from,
                to=whatsapp_to,
                body=message_body,
            )
        else:
            # Fallback to SMS
            sms_from = settings.TWILIO_PHONE_NUMBER
            message = client.messages.create(
                from_=sms_from,
                to=customer_number,
                body=message_body,
            )

        return True
    except Exception as e:
        print(f"SMS error: {e}")
        return False

def update_excel_sheet(inquiry):
    """Update Excel sheet with new inquiry data"""
    try:
        excel_file = os.path.join(settings.BASE_DIR, 'inquiries_data.xlsx')
        
        # Create new workbook if file doesn't exist
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Inquiries"
            # Add headers
            headers = ['Inquiry ID', 'Date', 'Name', 'Email', 'Phone', 'Company', 'Selected Plan', 'Message']
            ws.append(headers)
            
            # Style headers
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        else:
            wb = load_workbook(excel_file)
            ws = wb.active
        
        # Add new inquiry data
        row_data = [
            inquiry.inquiry_id,
            inquiry.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            inquiry.name,
            inquiry.email,
            inquiry.phone,
            inquiry.company or 'N/A',
            inquiry.selected_plan,
            inquiry.message or 'N/A'
        ]
        ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_file)
        return True
    except Exception as e:
        print(f"Excel error: {e}")
        return False
